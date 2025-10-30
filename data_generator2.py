import random
from datetime import datetime, timedelta
from openpyxl import load_workbook

# ------------------- CONFIG -------------------
template_path = "REFORM_Data_Template_FromSheet2.xlsx"
output_path = "REFORM_Data_DummyData_Region10_v2.xlsx"
num_rows = 7000

# ------------------- LOAD TEMPLATE -------------------
wb = load_workbook(template_path)
ws = wb["All_Columns"]

# ------------------- HELPER FUNCTIONS -------------------
def random_date(start_year=1950, end_year=2005):
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 12, 31)
    return (start + timedelta(days=random.randint(0, (end - start).days))).date()

def random_release_date(commit_date):
    delta = timedelta(days=random.randint(180, 3650))
    return commit_date + delta

def random_choice(options):
    return random.choice(options)

def random_bool():
    return random.choice([0, 1])

# ------------------- REGION 10 / CAGAYAN DE ORO SETTINGS -------------------
regions = ["Region X - Northern Mindanao"]
provinces = [
    "Misamis Oriental", "Misamis Occidental", "Bukidnon", "Lanao del Norte", "Camiguin"
]
cities = [
    "Cagayan de Oro City", "Iligan City", "Malaybalay City", "Valencia City",
    "Gingoog City", "Oroquieta City", "Ozamiz City", "Tangub City"
]
barangays_cdo = [
    "Barangay Carmen", "Barangay Kauswagan", "Barangay Balulang",
    "Barangay Lapasan", "Barangay Gusa", "Barangay Nazareth",
    "Barangay Bulua", "Barangay Lumbia", "Barangay Consolacion"
]
jails = [
    "CDO City Jail - Male Dorm", "CDO City Jail - Female Dorm",
    "Misamis Oriental Provincial Jail", "Bukidnon Provincial Jail",
    "Iligan City Jail", "Malaybalay City Jail"
]

# ------------------- DROPDOWN LISTS (MATCHING TEMPLATE) -------------------
SexList = ["Male", "Female"]
CivilStatusList = ["Single", "Married", "Separated", "Widowed", "Annulled", "Divorce"]
EducationList = ["None", "Elementary", "HighSchool", "College", "Vocational", "Graduate"]
ReligionList = ["Roman Catholic", "Islam", "Iglesia ni Cristo", "Born Again", "None"]
EmploymentList = ["Employed", "Unemployed", "Student", "Self-employed"]
CaseTypeList = [
    "Theft", "Robbery", "Qualified Theft", "Estafa", "Homicide", "Murder",
    "Physical Injuries", "Rape", "Illegal Drugs", "Illegal Possession of Firearm",
    "Kidnapping", "Arson", "Falsification", "Libel", "Direct Assault",
    "Violence Against Women and Their Children", "Child Abuse", "Fencing",
    "Graft and Corruption", "Carnapping"
]
CaseSeverityList = ["Minor", "Less Serious", "Serious"]
ReleaseOutcomeList = [
    "Acquitted", "Dismissed", "Bail", "Time Served", "Parole",
    "Convicted", "Probation", "Released", "Transferred", "Escaped", "Deceased"
]
SeasonalityList = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
AttendanceList = ["Rare", "Occasional", "Regular", "Consistent"]
JailLocationTypeList = ["Urban", "Rural"]

# ------------------- GENERATE 7,000 ROWS -------------------
print(f"🔧 Generating {num_rows} dummy rows for Region X...")

for i in range(2, num_rows + 2):  # row 2 to 7001
    # --- Target Labels ---
    ws[f"A{i}"] = random_bool()  # is_recidivist
    ws[f"B{i}"] = random_bool()  # is_habitual_delinquent
    ws[f"C{i}"] = random_bool()  # is_quasi_recidivist
    ws[f"D{i}"] = random_bool()  # is_reiteration
    ws[f"E{i}"] = random_bool()  # is_recommitted
    ws[f"F{i}"] = random_bool()  # is_no_rebooking

    # --- Profile Variables ---
    birthdate = random_date()
    age_at_commit = random.randint(18, 60)
    commit_date = datetime(birthdate.year + age_at_commit, random.randint(1, 12), random.randint(1, 28))
    age_at_release = age_at_commit + random.randint(0, 10)

    ws[f"G{i}"] = age_at_commit
    ws[f"H{i}"] = birthdate
    ws[f"I{i}"] = age_at_release
    ws[f"J{i}"] = random_choice(SexList)
    ws[f"K{i}"] = random_choice(["Carpentry", "Cooking", "None", "Farming", "Driving", "Tailoring"])
    ws[f"L{i}"] = random_choice(["Fair", "Dark", "Light"])
    ws[f"M{i}"] = random_choice(["Tattooed", "None"])
    ws[f"N{i}"] = random_choice(CivilStatusList)
    ws[f"O{i}"] = random.randint(0, 6)
    ws[f"P{i}"] = random_choice(EducationList)
    ws[f"Q{i}"] = random_choice(["Public School", "Private School", "None"])
    ws[f"R{i}"] = random_choice(ReligionList)
    ws[f"S{i}"] = random_choice(EmploymentList)
    ws[f"T{i}"] = random_choice(["Farmer", "Driver", "Student", "Vendor", "Laborer", "None"])

    # --- Residence ---
    ws[f"U{i}"] = random_choice(barangays_cdo)
    ws[f"V{i}"] = random_choice(cities)
    ws[f"W{i}"] = random_choice(provinces)
    ws[f"X{i}"] = random_choice(regions)
    ws[f"Y{i}"] = round(random.uniform(10, 40), 2)  # poverty_incidence
    ws[f"Z{i}"] = round(random.uniform(5, 25), 2)   # crime_rate_home
    ws[f"AA{i}"] = random_choice(["PWD", "LGBTQ", "Elderly", "Youth", "None"])

    # --- Case Variables ---
    ws[f"AB{i}"] = random_choice(CaseTypeList)
    ws[f"AC{i}"] = random_choice(CaseSeverityList)
    ws[f"AD{i}"] = random.randint(0, 5)
    ws[f"AE{i}"] = random_choice(ReleaseOutcomeList)
    ws[f"AF{i}"] = random.randint(10, 3650)
    ws[f"AG{i}"] = random_choice(SeasonalityList)
    ws[f"AH{i}"] = random_choice(barangays_cdo)
    ws[f"AI{i}"] = round(random.uniform(10, 40), 2)
    ws[f"AJ{i}"] = round(random.uniform(5, 25), 2)
    ws[f"AK{i}"] = random.randint(1, 31)
    ws[f"AL{i}"] = random_choice(SeasonalityList)
    ws[f"AM{i}"] = random.randint(2000, 2024)
    ws[f"AN{i}"] = random.randint(0, 365)
    ws[f"AO{i}"] = random.randint(100, 5000)
    ws[f"AP{i}"] = random_choice(["Yes", "No"])

    # --- Detention Experience ---
    # da_* indicators (0 or 1)
    da_flags = [random_bool() for _ in range(5)]
    ws[f"AR{i}"], ws[f"AS{i}"], ws[f"AT{i}"], ws[f"AU{i}"], ws[f"AV{i}"] = da_flags

    # program_* indicators (0 or 1)
    prog_flags = [random_bool() for _ in range(5)]
    ws[f"AX{i}"], ws[f"AY{i}"], ws[f"AZ{i}"], ws[f"BA{i}"], ws[f"BB{i}"] = prog_flags

    ws[f"BC{i}"] = random_choice(AttendanceList)
    ws[f"BD{i}"] = random_choice(["None", "Monthly", "Weekly"])
    ws[f"BE{i}"] = random_choice(JailLocationTypeList)
    ws[f"BF{i}"] = random_choice(jails)
    ws[f"BG{i}"] = random.randint(1, 5)
    ws[f"BH{i}"] = random.randint(0, 2000)
    ws[f"BI{i}"] = random_choice(CaseTypeList)

# ------------------- SAVE FILE -------------------
wb.save(output_path)
print(f"✅ Dummy data successfully generated and saved as: {output_path}")
print(f"📍 Focus Area: Cagayan de Oro City, Misamis Oriental, Region X - Northern Mindanao")
