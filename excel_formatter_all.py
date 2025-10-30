from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Create workbook
wb = Workbook()

# Create sheets - All_Columns will be the base sheet
all_columns_ws = wb.active
all_columns_ws.title = "All_Columns"
ws1 = wb.create_sheet("Target_Labels")
ws2 = wb.create_sheet("Profile_Variables")
ws3 = wb.create_sheet("Case_Variables")
ws4 = wb.create_sheet("Detention_Experience")
dropdown_ws = wb.create_sheet("DropdownLists")

# ------------------- STYLE DEFINITIONS -------------------
# Light blue fill for headers
header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# Bold font for headers
header_font = Font(bold=True, size=12)

# Border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Center alignment
center_align = Alignment(horizontal='center', vertical='center')

# ------------------- DROPDOWN LISTS SHEET -------------------
# Column headers for each dropdown
dropdown_lists = {
    "A": ("SexList", ["Male", "Female"]),
    "B": ("CivilStatusList", ["Single", "Married", "Separated", "Widowed", "Annulled", "Divorce"]),
    "C": ("EducationList", ["None", "Elementary", "HighSchool", "College", "Vocational", "Graduate"]),
    "D": ("ReligionList", ["Roman Catholic", "Islam", "Iglesia ni Cristo", "Born Again", "None"]),
    "E": ("EmploymentList", ["Employed", "Unemployed", "Student", "Self-employed"]),
    "F": ("CaseTypeList", [
        "Theft", "Robbery", "Qualified Theft", "Estafa", "Homicide", "Murder",
        "Physical Injuries", "Rape", "Illegal Drugs", "Illegal Possession of Firearm",
        "Kidnapping", "Arson", "Falsification", "Libel", "Direct Assault",
        "Violence Against Women and Their Children", "Child Abuse", "Fencing",
        "Graft and Corruption", "Carnapping"
    ]),
    "G": ("CaseSeverityList", ["Minor", "Less Serious", "Serious"]),
    "H": ("ReleaseOutcomeList", [
        "Acquitted", "Dismissed", "Bail", "Time Served", "Parole",
        "Convicted", "Probation", "Released", "Transferred", "Escaped", "Deceased"
    ]),
    "I": ("SeasonalityList", ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]),
    "J": ("AttendanceList", ["Rare", "Occasional", "Regular", "Consistent"]),
    "K": ("JailLocationTypeList", ["Urban", "Rural"]),
}

# Write dropdown data to "DropdownLists" sheet with styling
for col, (header, values) in dropdown_lists.items():
    cell = dropdown_ws[f"{col}1"]
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align
    
    for i, v in enumerate(values, start=2):
        cell = dropdown_ws[f"{col}{i}"]
        cell.value = v
        cell.border = thin_border

# Freeze header row in DropdownLists
dropdown_ws.freeze_panes = "A2"

# ------------------- ALL COLUMNS SHEET -------------------
# Define all columns from all sheets
all_columns = [
    # Target Labels
    "is_recidivist", "is_habitual_delinquent", "is_quasi_recidivist",
    "is_reiteration", "is_recommitted", "is_no_rebooking",
    
    # Profile Variables
    "age_at_commit", "birthdate", "age_at_release", "sex", "skills", "complexion",
    "tattoo_marks", "civil_status", "number_of_children", "education",
    "name_of_school", "religion", "employment_status", "actual_work",
    "residence_barangay", "residence_city_municipality", "residence_province",
    "residence_region", "poverty_incidence", "crime_rate_home", "vulnerable_sector",
    
    # Case Variables
    "case_type", "case_severity", "concurrent_cases", "release_outcome",
    "length_of_stay_days", "seasonality_arrest", "crime_location",
    "poverty_incidence_crime_area", "crime_rate_area", "crime_day", "crime_month",
    "crime_year", "reduced_sentence_days", "sentence_length_days", "parole_release",
    
    # Detention Experience
    "disciplinary_actions", "da_violence", "da_contraband", "da_disobedience",
    "da_escape_attempt", "da_property_damage", "program_participation",
    "prog_education", "prog_livelihood", "prog_counseling", "prog_spiritual",
    "prog_sports", "program_attendance", "visitor_frequency", "jail_location_type",
    "jail_location", "number_bookings", "days_since_last_release", "prior_offense_type"
]

# Write headers to All_Columns sheet with styling
for col_idx, header in enumerate(all_columns, 1):
    col_letter = get_column_letter(col_idx)
    cell = all_columns_ws[f"{col_letter}1"]
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align

# Map column names to their positions in All_Columns
column_positions = {col: idx + 1 for idx, col in enumerate(all_columns)}

# Helper function to add dropdown referencing DropdownLists
def add_dropdown(ws, col_letter, list_col_letter, start_row=2, end_row=500):
    formula = f"=DropdownLists!${list_col_letter}$2:${list_col_letter}$100"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

# Helper function to add date validation in YYYY-MM-DD format
def add_date_validation(ws, col_letter, start_row=2, end_row=500):
    # Create validation for date format YYYY-MM-DD
    dv = DataValidation(
        type="custom", 
        formula1=f'=AND(ISDATE({col_letter}{start_row}),LEN({col_letter}{start_row})=10,LEFT({col_letter}{start_row},4)>="1900",LEFT({col_letter}{start_row},4)<=YEAR(TODAY()))',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Date Format",
        error="Please enter date in YYYY-MM-DD format (e.g., 1990-05-15)"
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

# ------------------- ADD VALIDATIONS TO ALL_COLUMNS SHEET -------------------
# Add dropdown validations to All_Columns sheet using proper column letters
add_dropdown(all_columns_ws, get_column_letter(column_positions["sex"]), "A")  # sex -> SexList
add_date_validation(all_columns_ws, get_column_letter(column_positions["birthdate"]))  # birthdate
add_dropdown(all_columns_ws, get_column_letter(column_positions["civil_status"]), "B")  # civil_status
add_dropdown(all_columns_ws, get_column_letter(column_positions["education"]), "C")  # education
add_dropdown(all_columns_ws, get_column_letter(column_positions["religion"]), "D")  # religion
add_dropdown(all_columns_ws, get_column_letter(column_positions["employment_status"]), "E")  # employment_status
add_dropdown(all_columns_ws, get_column_letter(column_positions["case_type"]), "F")  # case_type
add_dropdown(all_columns_ws, get_column_letter(column_positions["case_severity"]), "G")  # case_severity
add_dropdown(all_columns_ws, get_column_letter(column_positions["release_outcome"]), "H")  # release_outcome
add_dropdown(all_columns_ws, get_column_letter(column_positions["seasonality_arrest"]), "I")  # seasonality_arrest
add_dropdown(all_columns_ws, get_column_letter(column_positions["program_attendance"]), "J")  # program_attendance
add_dropdown(all_columns_ws, get_column_letter(column_positions["jail_location_type"]), "K")  # jail_location_type

# ------------------- ADD COUNTING FORMULAS TO ALL_COLUMNS SHEET -------------------
# Function to add counting formulas for disciplinary_actions and program_participation
def add_counting_formulas(ws, start_row=2, end_row=500):
    # Get column letters for the counting columns
    disciplinary_col = get_column_letter(column_positions["disciplinary_actions"])
    program_part_col = get_column_letter(column_positions["program_participation"])
    
    # Get column letters for the sub-columns to count
    da_violence_col = get_column_letter(column_positions["da_violence"])
    da_contraband_col = get_column_letter(column_positions["da_contraband"])
    da_disobedience_col = get_column_letter(column_positions["da_disobedience"])
    da_escape_attempt_col = get_column_letter(column_positions["da_escape_attempt"])
    da_property_damage_col = get_column_letter(column_positions["da_property_damage"])
    
    prog_education_col = get_column_letter(column_positions["prog_education"])
    prog_livelihood_col = get_column_letter(column_positions["prog_livelihood"])
    prog_counseling_col = get_column_letter(column_positions["prog_counseling"])
    prog_spiritual_col = get_column_letter(column_positions["prog_spiritual"])
    prog_sports_col = get_column_letter(column_positions["prog_sports"])
    
    for row in range(start_row, end_row + 1):
        # Formula for disciplinary_actions: count how many 1s in da_* columns
        disciplinary_formula = f'=SUM({da_violence_col}{row}, {da_contraband_col}{row}, {da_disobedience_col}{row}, {da_escape_attempt_col}{row}, {da_property_damage_col}{row})'
        
        # Formula for program_participation: count how many 1s in prog_* columns
        program_formula = f'=SUM({prog_education_col}{row}, {prog_livelihood_col}{row}, {prog_counseling_col}{row}, {prog_spiritual_col}{row}, {prog_sports_col}{row})'
        
        # Apply formulas to cells
        ws[f'{disciplinary_col}{row}'] = disciplinary_formula
        ws[f'{program_part_col}{row}'] = program_formula

# Add the counting formulas
add_counting_formulas(all_columns_ws)

# Freeze header row in All_Columns
all_columns_ws.freeze_panes = "A2"

# ------------------- DERIVED SHEETS -------------------
# Function to style headers for derived sheets
def style_derived_sheet_headers(ws, column_names):
    for col_idx, header in enumerate(column_names, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

# Sheet 1: Target Labels
target_labels_columns = [
    "is_recidivist", "is_habitual_delinquent", "is_quasi_recidivist",
    "is_reiteration", "is_recommitted", "is_no_rebooking"
]
style_derived_sheet_headers(ws1, target_labels_columns)

# Sheet 2: Profile Variables
profile_variables_columns = [
    "age_at_commit", "birthdate", "age_at_release", "sex", "skills", "complexion",
    "tattoo_marks", "civil_status", "number_of_children", "education",
    "name_of_school", "religion", "employment_status", "actual_work",
    "residence_barangay", "residence_city_municipality", "residence_province",
    "residence_region", "poverty_incidence", "crime_rate_home", "vulnerable_sector"
]
style_derived_sheet_headers(ws2, profile_variables_columns)

# Sheet 3: Case Variables
case_variables_columns = [
    "case_type", "case_severity", "concurrent_cases", "release_outcome",
    "length_of_stay_days", "seasonality_arrest", "crime_location",
    "poverty_incidence_crime_area", "crime_rate_area", "crime_day", "crime_month",
    "crime_year", "reduced_sentence_days", "sentence_length_days", "parole_release"
]
style_derived_sheet_headers(ws3, case_variables_columns)

# Sheet 4: Detention Experience
detention_experience_columns = [
    "disciplinary_actions", "da_violence", "da_contraband", "da_disobedience",
    "da_escape_attempt", "da_property_damage", "program_participation",
    "prog_education", "prog_livelihood", "prog_counseling", "prog_spiritual",
    "prog_sports", "program_attendance", "visitor_frequency", "jail_location_type",
    "jail_location", "number_bookings", "days_since_last_release", "prior_offense_type"
]
style_derived_sheet_headers(ws4, detention_experience_columns)

# ------------------- ADD FORMULAS TO DERIVED SHEETS -------------------
# Function to add formulas that reference All_Columns sheet (with blank when empty)
def add_formulas_to_derived_sheet(derived_ws, column_names, start_row=2, end_row=500):
    for row in range(start_row, end_row + 1):
        for col_idx, col_name in enumerate(column_names, start=1):
            if col_name in column_positions:
                all_col_letter = get_column_letter(column_positions[col_name])
                cell = derived_ws.cell(row=row, column=col_idx)
                # Use IF formula to show blank when source is empty/0
                cell.value = f'=IF(ISBLANK(All_Columns!{all_col_letter}{row}),"",All_Columns!{all_col_letter}{row})'
                # Add border to data cells
                cell.border = thin_border

# Add formulas to all derived sheets
add_formulas_to_derived_sheet(ws1, target_labels_columns)
add_formulas_to_derived_sheet(ws2, profile_variables_columns)
add_formulas_to_derived_sheet(ws3, case_variables_columns)
add_formulas_to_derived_sheet(ws4, detention_experience_columns)

# Freeze header rows in all derived sheets
ws1.freeze_panes = "A2"
ws2.freeze_panes = "A2"
ws3.freeze_panes = "A2"
ws4.freeze_panes = "A2"

# ------------------- SAVE -------------------
wb.save("REFORM_Data_Template_AllSheet2.xlsx")
print("✅ Excel created successfully: REFORM_Data_Template_FromSheet.xlsx")
print("📊 Structure:")
print("   - All_Columns: Base sheet with all data and validations")
print("   - Other sheets: Derived views with formulas referencing All_Columns")
print("   - DropdownLists: Source data for dropdown validations")
print("🎨 Styling Applied:")
print("   - Light blue headers with bold, larger font")
print("   - 1pt borders on all cells")
print("   - Centered header text")
print("   - Frozen header rows")
print("🔧 Formula Features:")
print("   - Derived sheets show blank instead of 0 when source is empty")
print("   - disciplinary_actions automatically counts 1s in da_* columns")
print("   - program_participation automatically counts 1s in prog_* columns")

# Print column mapping for reference
print("\n📋 Column positions in All_Columns sheet:")
for i, col in enumerate(all_columns, 1):
    print(f"   {get_column_letter(i)}: {col}")

print("\n🔢 Automatic Counting:")
print("   - disciplinary_actions = SUM(da_violence, da_contraband, da_disobedience, da_escape_attempt, da_property_damage)")
print("   - program_participation = SUM(prog_education, prog_livelihood, prog_counseling, prog_spiritual, prog_sports)")